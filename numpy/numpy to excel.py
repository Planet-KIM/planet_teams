#엑셀을 만들고 불러와서 0,1로 단순화 그리고 그 그래프를 엑셀 원래 표 옆에다 지정해서 배치
import numpy as np
import pandas as pd
from openpyxl import Workbook
import openpyxl

A = np.random.randn(5,5)
graph = pd.DataFrame(A)
print(graph)

#graph.to_excel(excel_writer ='ntop.xlsx')

open = openpyxl.load_workbook('ntop.xlsx', data_only=True)
active = open.active

count = 0
sheet = open.worksheets[0]
for item in sheet.rows:
    print(item[count].value)
    count+=1

counta = 0
for i in open:
    countb = 0

    for j in i:
        if 0 > open[counta][countb]:
            open[counta][countb] = 0

        else:
            open[counta][countb] = 1

        countb = countb + 1
    counta = counta + 1


write_wb = Workbook()
write_wb.cell(7,7,open)
write_wb.save('test.xlsx')







