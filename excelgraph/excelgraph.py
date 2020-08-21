from pandas import ExcelWriter
from openpyxl import load_workbook
import openpyxl
import pandas as pd
import numpy as np

content = np.random.randn(10, 6)
graph = pd.DataFrame(content, columns = ['5', '6', '7', '8', '9', '10'])
print(graph)

graph.to_excel(excel_writer ='graph.xlsx')
#wb = openpyxl.Workbook()
#wb.save('graph.xlsx')

filename = 'graph.xlsx'
wb = load_workbook(filename)
ws = wb.active

content = np.random.randn(10, 6)
#graph = pd.DataFrame(content, columns = ['5', '6', '7', '8', '9', '10'])

counta = 0
for i in content:
    countb= 0
    for j in i:

        if 0 < content[counta][countb]:
            content[counta][countb] = 1

        else:
            content[counta][countb] = 0

        countb = countb + 1
    counta = counta + 1

A = pd.DataFrame(content, columns = ['11', '12', '13', '14', '15', '17'])
#ws['H1'] = print(A)
#wb.save('graph.xlsx')



